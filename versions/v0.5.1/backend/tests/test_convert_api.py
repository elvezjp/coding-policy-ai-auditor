"""変換APIエンドポイントのテスト"""

import pytest
from io import BytesIO
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


@pytest.fixture
def client():
    """テストクライアント"""
    return TestClient(app)


def create_test_excel(data: list[list], sheet_name: str = "Sheet1") -> bytes:
    """テスト用Excelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class TestAuditorExcelSpreadsheetEndpoint:
    """POST /api/convert/auditor-excel-spreadsheet エンドポイントのテスト"""

    def test_valid_spreadsheet(self, client):
        """正常なスプレッドシート形式の解析"""
        data = [
            ["A1", "B1", "C1"],
            ["A2", "B2", "C2"],
        ]
        content = create_test_excel(data)

        response = client.post(
            "/api/convert/auditor-excel-spreadsheet",
            files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        result = response.json()
        assert result["success"] is True
        assert result["filename"] == "test.xlsx"
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["title"] == "Sheet1"
        assert result["sheets"][0]["rows"] == 2
        assert result["sheets"][0]["columns"] == 3

    def test_multiple_sheets(self, client):
        """複数シートの解析"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "シート1"
        ws1.append(["データ1"])

        ws2 = wb.create_sheet("シート2")
        ws2.append(["データ2"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        response = client.post(
            "/api/convert/auditor-excel-spreadsheet",
            files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        result = response.json()
        assert result["success"] is True
        assert len(result["sheets"]) == 2
        assert result["sheets"][0]["title"] == "シート1"
        assert result["sheets"][1]["title"] == "シート2"

    def test_invalid_file_format(self, client):
        """無効なファイル形式"""
        response = client.post(
            "/api/convert/auditor-excel-spreadsheet",
            files={"file": ("test.txt", b"invalid content", "text/plain")},
        )

        assert response.status_code == 200
        result = response.json()
        assert result["success"] is False
        assert "対応していないファイル形式" in result["error"]
