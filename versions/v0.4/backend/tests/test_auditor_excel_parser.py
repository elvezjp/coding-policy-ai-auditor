"""AIオーディター形式Excel解析サービスのテスト"""

from io import BytesIO
from openpyxl import Workbook

from app.services.auditor_excel_parser import (
    parse_auditor_excel_spreadsheet,
    _make_sheet_key,
    _parse_hyperlink_location,
)


def create_test_excel(data: list[list], sheet_name: str = "Sheet1") -> bytes:
    """テスト用Excelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class TestParseAuditorExcelSpreadsheet:
    """parse_auditor_excel_spreadsheet関数のテスト"""

    def test_single_sheet(self):
        """単一シートの解析"""
        data = [
            ["A1", "B1", "C1"],
            ["A2", "B2", "C2"],
        ]
        content = create_test_excel(data)

        result = parse_auditor_excel_spreadsheet(content, "test.xlsx")

        assert result["success"] is True
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["title"] == "Sheet1"
        assert result["sheets"][0]["rows"] == 2
        assert result["sheets"][0]["columns"] == 3
        assert result["sheets"][0]["data"][0][0] == "A1"

    def test_multiple_sheets(self):
        """複数シートの解析"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "シート1"
        ws1.append(["データ1"])

        ws2 = wb.create_sheet("シート2")
        ws2.append(["データ2"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        result = parse_auditor_excel_spreadsheet(content, "test.xlsx")

        assert result["success"] is True
        assert len(result["sheets"]) == 2
        assert result["sheets"][0]["title"] == "シート1"
        assert result["sheets"][1]["title"] == "シート2"

    def test_sheet_key_generation(self):
        """シート識別キーの生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "テスト (シート) 1"
        ws.append(["データ"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        result = parse_auditor_excel_spreadsheet(content, "test.xlsx")

        assert result["success"] is True
        # 特殊文字が "_" に置換される
        assert result["sheets"][0]["key"] == "テスト__シート__1"

    def test_empty_cells(self):
        """空セルの処理"""
        data = [
            ["A1", None, "C1"],
            [None, "B2", None],
        ]
        content = create_test_excel(data)

        result = parse_auditor_excel_spreadsheet(content, "test.xlsx")

        assert result["success"] is True
        assert result["sheets"][0]["data"][0][0] == "A1"
        assert result["sheets"][0]["data"][0][1] == ""  # Noneは空文字列に
        assert result["sheets"][0]["data"][1][0] == ""


class TestHelperFunctions:
    """ヘルパー関数のテスト"""

    def test_make_sheet_key(self):
        """シート名からキー生成"""
        assert _make_sheet_key("シート1") == "シート1"
        assert _make_sheet_key("Sheet (1)") == "Sheet__1_"
        assert _make_sheet_key("テスト-データ") == "テスト_データ"

    def test_parse_hyperlink_location_with_quotes(self):
        """シングルクォート付きハイパーリンクの解析"""
        result = _parse_hyperlink_location("'シート名'!A1")
        assert result == ("シート名", "A1")

    def test_parse_hyperlink_location_without_quotes(self):
        """シングルクォートなしハイパーリンクの解析"""
        result = _parse_hyperlink_location("シート名!B2")
        assert result == ("シート名", "B2")

    def test_parse_hyperlink_location_invalid(self):
        """無効なハイパーリンクの処理"""
        assert _parse_hyperlink_location(None) is None
        assert _parse_hyperlink_location("") is None
        assert _parse_hyperlink_location("no_exclamation") is None
