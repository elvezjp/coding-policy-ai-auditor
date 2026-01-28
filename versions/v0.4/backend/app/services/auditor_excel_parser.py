"""AIオーディター形式Excel解析サービス"""

import re
from io import BytesIO
from typing import TypedDict

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


class CellHyperlink(TypedDict):
    """セルのハイパーリンク情報"""
    sheetName: str      # リンク先シート名
    cell: str           # リンク先セル（例: "A1"）


class SpreadsheetSheet(TypedDict):
    """スプレッドシートのシートデータ"""
    title: str          # シート名
    key: str            # 識別キー
    data: list[list[str]]  # 2次元配列データ
    rows: int           # 行数
    columns: int        # 列数
    hyperlinks: dict[str, CellHyperlink]  # ハイパーリンク情報 {"row,col": {sheetName, cell}}


class SpreadsheetParseResult(TypedDict):
    """スプレッドシート形式の解析結果"""
    success: bool
    sheets: list[SpreadsheetSheet]
    error: str | None


def _make_sheet_key(title: str) -> str:
    """シート名から識別キーを生成する"""
    # 特殊文字を置換してキーを生成
    key = re.sub(r'[^a-zA-Z0-9\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fff]', '_', title)
    return key


def _parse_hyperlink_location(location: str | None) -> tuple[str, str] | None:
    """
    ハイパーリンクのlocationからシート名とセルを抽出する

    例: "'全般(フォーマット)1'!A1" -> ("全般(フォーマット)1", "A1")
    例: "実装コメント1!A1" -> ("実装コメント1", "A1")

    Args:
        location: ハイパーリンクのlocation文字列

    Returns:
        (シート名, セル) のタプル、または解析できない場合はNone
    """
    if not location:
        return None

    # "!"でシート名とセルを分割
    if '!' not in location:
        return None

    parts = location.rsplit('!', 1)
    if len(parts) != 2:
        return None

    sheet_name, cell = parts

    # シート名のシングルクォートを除去
    if sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1]

    return (sheet_name, cell)


def _get_sheet_range(ws) -> tuple[int, int, int, int]:
    """
    シートの表示範囲を取得する

    1. まず印刷範囲（print_area）を確認
    2. 印刷範囲がない場合は、データが存在する範囲（used range）を使用

    Args:
        ws: Worksheetオブジェクト

    Returns:
        tuple: (min_row, min_col, max_row, max_col) 1-based
    """
    sheet_max_row = ws.max_row or 1
    sheet_max_col = ws.max_column or 1

    # 印刷範囲を確認
    try:
        pa = ws.print_area
        if pa:
            # Print_Areaは文字列（単一範囲）または反復可能（複数範囲）
            if isinstance(pa, str):
                ranges = [pa]
            else:
                try:
                    ranges = list(pa)
                except (TypeError, AttributeError):
                    ranges = [pa]

            # 最初の印刷範囲を使用
            for r in ranges:
                try:
                    range_str = str(r)
                    # シート名を除去（例: "'Sheet1'!$A$1:$Z$100" -> "$A$1:$Z$100"）
                    if '!' in range_str:
                        range_str = range_str.split('!', 1)[1]

                    min_col, min_row, max_col, max_row = range_boundaries(range_str)

                    # 範囲の検証
                    if min_row > max_row or min_col > max_col:
                        continue
                    if min_row < 1 or min_col < 1:
                        continue

                    # シートの最大範囲を超えないように制限
                    max_row = min(max_row, sheet_max_row)
                    max_col = min(max_col, sheet_max_col)

                    return (min_row, min_col, max_row, max_col)
                except Exception:
                    continue
    except Exception:
        pass

    # 印刷範囲がない場合、データが存在する範囲を使用
    try:
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        return (min_row, min_col, max_row, max_col)
    except Exception:
        return (1, 1, sheet_max_row, sheet_max_col)


def parse_auditor_excel_spreadsheet(content: bytes, filename: str) -> SpreadsheetParseResult:
    """
    AIオーディター形式のExcelファイルをスプレッドシート形式で解析する

    全シートを2次元配列形式で返す。Tabulatorのスプレッドシートモードで
    そのまま表示できる形式。

    Args:
        content: Excelファイルのバイナリデータ
        filename: ファイル名（エラーメッセージ用）

    Returns:
        SpreadsheetParseResult: 解析結果（全シートの2次元配列データ）
    """
    try:
        # Excelファイルを読み込み（read_only=Falseでハイパーリンク情報も取得可能）
        wb = load_workbook(filename=BytesIO(content), read_only=False, data_only=True)

        if not wb.sheetnames:
            return {
                "success": False,
                "sheets": [],
                "error": "シートが見つかりません",
            }

        sheets: list[SpreadsheetSheet] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data: list[list[str]] = []
            hyperlinks: dict[str, CellHyperlink] = {}

            # 印刷範囲またはデータ範囲を取得
            min_row, min_col, max_row, max_col = _get_sheet_range(ws)

            # 指定範囲のみを2次元配列として取得（values_only=Falseでセルオブジェクトを取得）
            for row_idx, row in enumerate(ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=False
            )):
                row_data: list[str] = []
                for col_idx, cell in enumerate(row):
                    # セルの値を取得
                    value = str(cell.value) if cell.value is not None else ""
                    row_data.append(value)

                    # ハイパーリンク情報を取得
                    if cell.hyperlink and cell.hyperlink.location:
                        parsed = _parse_hyperlink_location(cell.hyperlink.location)
                        if parsed:
                            sheet_target, cell_target = parsed
                            # キーは "行,列" 形式（0-based）
                            hyperlinks[f"{row_idx},{col_idx}"] = {
                                "sheetName": sheet_target,
                                "cell": cell_target,
                            }

                data.append(row_data)

            # 行数と列数を計算
            num_rows = len(data)
            num_cols = max((len(row) for row in data), default=0)

            sheets.append({
                "title": sheet_name,
                "key": _make_sheet_key(sheet_name),
                "data": data,
                "rows": num_rows,
                "columns": num_cols,
                "hyperlinks": hyperlinks,
            })

        wb.close()

        return {
            "success": True,
            "sheets": sheets,
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "sheets": [],
            "error": f"Excel解析エラー: {str(e)}",
        }
